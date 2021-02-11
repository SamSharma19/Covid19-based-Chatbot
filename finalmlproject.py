import random
import datetime

'''for reading and writting in the excel file'''
from openpyxl import load_workbook
wb = load_workbook('COVID.xlsx')
sheet = wb.active

'''function to extract last character of word and add it as dictionary'''
def extraction(word):
    return{'lastchar' : word[-1]}

'''list of names extracted from a file that are stored along with their gender'''
from nltk.corpus import names
malenames = [ (name , 'male') for name in names.words('male.txt') ]
femalenames = [ (name, 'female') for name in names.words('female.txt')]
allnames = malenames + femalenames
random.shuffle(allnames)

'using naive bayes classifier to detect the name is male or female'
from nltk import NaiveBayesClassifier as nbc

'''we will use last character and gender of the names as the features'''
total = [ (extraction(n),gender) for (n,gender) in allnames]
training_set = total[:1000]
test_set = total[1000:]

'''training the model'''
classifier = nbc.train(training_set)


print('Hello, can you please tell me your name?')
name = input()
m = 0
f = 0
s = classifier.classify(extraction(name))
if(s == 'male'):
    name = 'Mr ' + name
    sex = 'M'
    m = 1
else:
    name = 'Ms ' + name
    sex = 'F'
    f = 1
    
print('Hello ' + name + ' I am EDITH and I perfectly designed to diagnose you for COVID-19 or chat.')
print('You are logged in at')

'''extra check for sex of person as sometimes male name can be pronounced as female vice-versa too'''
print(name + ' are you a ' + s + '?')
rr = input()
if 'no' in rr.lower():
    print('Please enter you correct sex')
    rs = input()
    if 'male' in rs.lower():
        sex = 'M'
        name = 'Mr ' + name[3:]
        m = 1
    else:
        sex = 'F'
        f = 1
        name = 'Ms ' + name[3:]

'''printing the current month, current year, current hour, current minute of the system'''
e = datetime.datetime.now()
print(e)
log = str(e.month) + str(e.year) + str(e.hour) + str(e.minute)
print(str(e.month) + '/' + str(e.year) + '--'+ str(e.hour) + ':' + str(e.minute))

print(name + ' , Do you want to have COVID-19 test or chat?')
ans = input()
   

if 'test' in ans.lower():
        '''questions to collect data to predict the status of person wether he has contracted corona or not'''
        
        print(name + ', Can you tell me your age')
        age = input()
        age = int(age)
        '''adding weight to ages according to the survey'''
        if(age<=17):
            age = 0.06
        elif(age>17 and age<=44):
            age = 3.9
        elif(age>44 and age<=64):
            age = 22.4
        elif(age>64 and age<=74):
            age = 24.9
        else:
            age = 48.7
       
        print('Are you Experiencing any of the following symptom')
        print(' Cough , Fever , Difficulty in Breathing  or none')
        symp = input()
        if symp.lower() == "fever" :
            symp = 83
        elif symp.lower() == "cough" :
            symp = 82
        elif 'brea' in symp.lower():
            symp = 31
        else:
            symp = 1
            
        print('Have you had any of the following disease before')
        print('Diabetes, Hypertension ,Lung Disease, Heart Disease or none of them')
        dise = input()
        if 'none' in dise.lower():
           dise = 1
           dise = int(dise)
        elif 'diabetes' in dise.lower():
            dise = 30
            dise = int(dise)
        elif 'lung' in dise.lower():
            
            dise = 70
            dise = int(dise)
        elif 'tension' in dise.lower():
            
            dise = 20
            dise = int(dise)
        else:
            print('You didint entered a valid option')
            s = False
    
        print('Have you travelled anywhere in the last 14 days(Yes or No)')
        travel = input()
        if 'y' in travel.lower():
            travel = 1
        else:
            travel = 0
        
        print("Which of the following applies to you?")
        print("I have recently interacted or lived with someone who has tested positive for COVID-19, I am a healthcare \
            worker and I examined a COVID-19 confirmed case without protective gear or None of above")
        env = input()
        if 'lived' in env.lower():
            env = 70
        elif 'worked' in env.lower():
            env = 50
        else:
            env = 0
        
        print('Can you tell in which state do you live')
        state = input()

        print('Thank you for your time '+ name + ' I will tell you your results shortly')
        reply  = [log, age, sex, symp, dise, travel, env, state]
        

        '''printing the entered data according multiplied by its weight accordingly'''       
        print(reply)

        '''finding empty row in the excel file'''
        for i in range(1,800):
            if (sheet.cell(row = i, column = 2).value == None):
                index = i
                break
        print(i)

        '''updating the sheet with entered values according to their weights and saving the file'''
        for j in range(0,8):
            sheet.cell(row = index, column = j+1).value = reply[j]
        print(log)
        wb.save('COVID.xlsx')

        '''--------------------------------------------------------------------------------------------'''
        '''Code to predict whether person is affected or not based on the features'''
        import numpy as np
        import matplotlib.pyplot as plt
        import pandas as pd

        '''reading the csv file that will be used to train the model'''
        dataset = pd.read_csv('Book1.csv')
        x = dataset.iloc[:,1:-2].values
        y = dataset.iloc[:,-1].values    

        '''encoding sex of the people(as is categorical variable)'''
        from sklearn.preprocessing import OneHotEncoder
        ohe = OneHotEncoder(sparse = False)
        z = ohe.fit_transform(x[:,1].reshape(-1,1))
        x = np.delete(x,1,axis = 1)
        x = np.append(x,z,axis = 1)

        '''splitting the dataset into test and train set'''
        from sklearn.model_selection import train_test_split
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size = 0.2, random_state = 0)

        '''feature scaling each column accordingly '''
        from sklearn.preprocessing import StandardScaler
        sc_age = StandardScaler()
        sc_symp = StandardScaler()
        sc_dise = StandardScaler()
        sc_travel = StandardScaler()
        sc_env = StandardScaler()
        x_age = sc_age.fit_transform(x_train[:,0].reshape(len(x_train[:,0]),1)) 
        x_symp = sc_symp.fit_transform(x_train[:,1].reshape(len(x_train[:,0]),1)) 
        x_dise = sc_dise.fit_transform(x_train[:,2].reshape(len(x_train[:,0]),1)) 
        x_travel = sc_travel.fit_transform(x_train[:,3].reshape(len(x_train[:,0]),1)) 
        x_env = sc_env.fit_transform(x_train[:,4].reshape(len(x_train[:,0]),1)) 
        training_set = np.concatenate((x_age,x_symp,x_dise,x_travel,x_env,x_train[:,5].reshape(len(x_train[:,5]),1),x_train[:,6].reshape(len(x_train[:,6]),1)), axis = 1)
        xx_age = sc_age.transform(x_test[:,0].reshape(len(x_test[:,0]),1))
        xx_symp = sc_age.transform(x_test[:,0].reshape(len(x_test[:,1]),1))
        xx_dise = sc_age.transform(x_test[:,0].reshape(len(x_test[:,2]),1))
        xx_travel = sc_age.transform(x_test[:,0].reshape(len(x_test[:,3]),1))
        xx_env = sc_age.transform(x_test[:,0].reshape(len(x_test[:,4]),1))
        test_set = np.concatenate((xx_age,xx_symp,xx_dise,xx_travel,xx_env,x_test[:,5].reshape(len(x_test[:,5]),1),x_test[:,6].reshape(len(x_test[:,6]),1)), axis = 1)

        '''classifying the dataset using svm algorithm'''
        from sklearn.svm import SVC
        classifier = SVC(kernel = 'rbf', random_state = 0)
        classifier.fit(training_set,y_train)

        '''feature scaling the entered data accordingly'''
        age = sc_age.transform([[age]])
        symp = sc_symp.transform([[symp]])
        dise = sc_age.transform([[dise]])
        travel = sc_age.transform([[travel]])
        env = sc_age.transform([[env]])

        '''predicted result wether person has corona or not'''
        result = classifier.predict([[age,symp,dise,travel,env,f,m]])

        if 'No' in result:
            pd = 0
            print('Congratulations ' + name + ' as per my analysis you are healthy')
        else:
            pd = 1
            print('Here are some references you can look')
            '''online consultation website'''
            from selenium import webdriver
            from selenium.webdriver.common.keys import Keys
            import os
            import time
            driver = webdriver.Chrome(r'"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"')
            driver.get("https://www.practo.com/consult/direct/new_consultation?utm_source=google&utm_medium=cpc&utm_\
                campaign=mweb-consult-sem-tier-2-generic&gclid=Cj0KCQjwwr32BRD4ARIsAAJNf_0k-tj0b3GqNzk9DQ4AcvYY5gpJn0ed59Xe21LflDukkAtdex0pxIIaApc_EALw_wcB")
            
        '''updating the predicted value in the excel sheet for the entered data'''
        wb = load_workbook(r'COVID.xlsx')
        sheet = wb.active
        if pd == 0:
            sheet.cell(row = index, column = 9).value = 'No'
        else:
            sheet.cell(row = index, column = 9).value = 'Yes'
        wb.save('COVID.xlsx')

        ypred = classifier.predict(test_set)

        from sklearn.metrics import confusion_matrix
        cm = confusion_matrix(y_test, ypred)

#------------- Snetiment Analysis --------------------------

elif 'chat' in ans.lower():

    from textblob import TextBlob
    '''assigning user nickname'''
    print('Hello ' + name + ' May I know your nickname?')
    nick = input()
    
    '''list of greeting to greet the user'''    
    greeting = [ 'How are you ?', 'Whats going on ? ', 'How is your day? ','Wassup? ']
    print(random.choice(greeting) + nick)
    reply = input()
    blob = TextBlob(reply)
    if (blob.polarity > 0):
        print('Glad to meet you I am ADSPARX - a chatbot provided by EDITH.')
    else:
        print('Sorry to bother you')
    
    topics = ['Did you hear about COVID ?', 'How is the weather over there?', 'What are you doing right now?']
    print(random.choice(topics))
    reply1 = input()
    blob1 = TextBlob(reply1)
    if(blob1.polarity > 0.5):
        print('Glad to hear that')
    elif blob.polarity > 0:
        print('Oh I see')
    else :
        print('Sorry to hear that')

    print("How is COVID situation in your locality?")
    reply2 = input()
    blob2 = TextBlob(reply2)
    if blob2.polarity > 0:
        print('Great, but do take precautions before leaving home.')
    else :
        print('Sorry to hear that. Stay safe.')

    print("Are you a student or employee?")
    you = input()
    if 'student' in you.lower():
        print("Has been any cases in your school/college?")
        ans1 = input()

        if 'y' in ans1.lower():
            print("Have any of your teacher/friend affected by COVID?")
            ans2 = input()
            if 'n' in ans2.lower():
                print('Glad to hear that.')
            else :
                print('Sorry to hear that. Hope, they get well soon.')

        print("Are classes being conducted online?")
        reply3 = input()
        blob3 = TextBlob(reply3)
        if(blob3.polarity > 0.5):
            print('Oh great!!')
        elif(blob3.polarity > 0):
            print('Okay')
        else :
            print('Hope, the classes are not too boring.')

    elif 'employee' in you.lower():
        print("Has been any cases in your workplace?")
        ans3 = input()

        if 'y' in ans3.lower():
            print("Have any of your friends affected by COVID?")
            ans4 = input()
            if 'n' in ans4.lower():
                print('Glad to hear that.')
            else :
                print('Sorry to hear that. Hope, they get well soon.')

        print("Are meetings being conducted online?")
        reply6 = input()
        blob6 = TextBlob(reply6)
        if(blob6.polarity > 0.5):
            print('Oh great!!')
        elif(blob6.polarity > 0):
            print('Okay')
        else :
            print('Hope, the meeting are not too boring.')

    if (blob.polarity > 0):
        print('I had a wonderful time chatting with you.')
    else:
        print('Sorry, for taking your time.')

                
else:
    print("Sorry, I haven't configured to perform such function.")
